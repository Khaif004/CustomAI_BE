import logging
import json
import csv
import io
from typing import Tuple

logger = logging.getLogger(__name__)

ALLOWED_EXTENSIONS = {
    ".pdf", ".docx", ".xlsx", ".xls", ".csv",
    ".json", ".txt", ".md", ".py", ".js", ".ts",
    ".java", ".html", ".css", ".xml", ".yaml", ".yml",
    ".sql", ".sh", ".bat", ".log", ".env", ".cfg", ".ini",
}

MAX_FILE_SIZE = 10 * 1024 * 1024


def get_file_extension(filename: str) -> str:
    return "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def validate_file(filename: str, size: int) -> Tuple[bool, str]:
    ext = get_file_extension(filename)
    if ext not in ALLOWED_EXTENSIONS:
        return False, f"Unsupported file type: {ext}. Supported: {', '.join(sorted(ALLOWED_EXTENSIONS))}"
    if size > MAX_FILE_SIZE:
        return False, f"File too large. Maximum size is {MAX_FILE_SIZE // (1024*1024)}MB"
    return True, ""


async def extract_text(filename: str, content: bytes) -> str:
    ext = get_file_extension(filename)

    try:
        if ext == ".pdf":
            return _parse_pdf(content)
        elif ext == ".docx":
            return _parse_docx(content)
        elif ext in (".xlsx", ".xls"):
            return _parse_excel(content)
        elif ext == ".csv":
            return _parse_csv(content)
        elif ext == ".json":
            return _parse_json(content)
        else:
            return content.decode("utf-8", errors="replace")
    except Exception as e:
        logger.error(f"Error parsing {filename}: {e}")
        raise ValueError(f"Failed to parse {filename}: {str(e)}")


def _parse_pdf(content: bytes) -> str:
    from PyPDF2 import PdfReader
    reader = PdfReader(io.BytesIO(content))
    pages = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text()
        if text and text.strip():
            pages.append(f"--- Page {i+1} ---\n{text.strip()}")
    return "\n\n".join(pages) if pages else "(No readable text found in PDF)"


def _parse_docx(content: bytes) -> str:
    from docx import Document
    doc = Document(io.BytesIO(content))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs) if paragraphs else "(No readable text found in document)"


def _parse_excel(content: bytes) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            row_vals = [str(c) if c is not None else "" for c in row]
            if any(v.strip() for v in row_vals):
                rows.append(" | ".join(row_vals))
        if rows:
            header = f"### Sheet: {sheet_name}\n"
            sheets.append(header + "\n".join(rows))
    wb.close()
    return "\n\n".join(sheets) if sheets else "(No data found in spreadsheet)"


def _parse_csv(content: bytes) -> str:
    text = content.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = []
    for row in reader:
        rows.append(" | ".join(row))
    return "\n".join(rows) if rows else "(Empty CSV file)"


def _parse_json(content: bytes) -> str:
    data = json.loads(content.decode("utf-8"))
    return json.dumps(data, indent=2, ensure_ascii=False)
